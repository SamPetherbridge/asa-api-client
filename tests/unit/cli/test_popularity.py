"""Tests for the Search Popularity enrichment (``cli.popularity``)."""

import json
import re
from collections.abc import Callable
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import pytest
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import ec
from openpyxl import load_workbook
from pytest_httpx import HTTPXMock
from typer.testing import CliRunner

from asa_api_client.cli import fetch, popularity
from asa_api_client.cli.analyze import app
from asa_api_client.v1.client import AppleAdsClient
from tests.unit.cli.conftest import (
    API,
    TOKEN_URL,
    campaigns_json,
    report_json,
    report_row,
    token_json,
)

V1 = "https://api.ads.apple.com/v1"
ACLS_URL = f"{V1}/acls"
POPULARITY_URL = f"{V1}/insights/apps/search-term-popularity/query"

TODAY = date(2026, 8, 17)  # a Monday; latest complete week is Aug 9 - Aug 15

runner = CliRunner()


@pytest.fixture(scope="module")
def pem() -> str:
    """Generate a throwaway EC P-256 private key in PEM format."""
    key = ec.generate_private_key(ec.SECP256R1())
    return key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption(),
    ).decode()


def _factory(pem: str, ad_account_id: str | None = "12345") -> Callable[[], AppleAdsClient]:
    """Build a client factory with dummy credentials for injection.

    Args:
        pem: The PEM-encoded EC private key.
        ad_account_id: The preselected ad account, or None to force
            ACL-based account resolution.

    Returns:
        A zero-argument callable producing a configured client.
    """
    return lambda: AppleAdsClient(
        client_id="SEARCHADS.test",
        team_id="TEAM123",
        key_id="KEY123",
        ad_account_id=ad_account_id,
        private_key=pem,
    )


def _fetch_result(keywords: list[str], search_terms: list[str]) -> fetch.FetchResult:
    """Build a FetchResult with keyword and search-term level frames.

    Args:
        keywords: Values for the keywords level's ``keyword`` column.
        search_terms: Values for the search_terms level's
            ``search_term_text`` column.

    Returns:
        A minimal FetchResult carrying just the two term-bearing levels.
    """
    kw_daily = pd.DataFrame({"keyword": keywords}) if keywords else pd.DataFrame()
    st_daily = pd.DataFrame({"search_term_text": search_terms}) if search_terms else pd.DataFrame()
    return fetch.FetchResult(
        levels={
            "keywords": fetch.LevelData(label="Keywords", daily=kw_daily),
            "search_terms": fetch.LevelData(label="Search Terms", daily=st_daily),
        },
        prior_campaigns=pd.DataFrame(),
    )


def _mock_token(httpx_mock: HTTPXMock) -> None:
    """Mock the OAuth token endpoint (reusable, optional)."""
    httpx_mock.add_response(url=TOKEN_URL, json=token_json(), is_optional=True, is_reusable=True)


def _popularity_row(
    term: str,
    *,
    country: str = "US",
    genre: str = "Games",
    rank: int = 1,
    p100: int = 50,
    p5: int = 3,
) -> dict[str, Any]:
    """Build one search-term-popularity API row.

    Args:
        term: The search term.
        country: The country/region code.
        genre: The App Store genre.
        rank: Rank within country + genre.
        p100: Popularity on the 1-100 scale.
        p5: Popularity on the 1-5 scale.

    Returns:
        The camelCase JSON row.
    """
    return {
        "week": "2026-08-09",
        "countryOrRegion": country,
        "genre": genre,
        "searchTerm": term,
        "rankInGenre": rank,
        "searchPopularityInGenre": p100,
        "searchPopularity1to100": p100,
        "searchPopularity1to5": p5,
    }


def _popularity_json(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Wrap popularity rows in the v1 insights response envelope."""
    return {"result": {"rows": rows}}


class TestLatestCompleteWeek:
    """Sun-Sat week alignment used for popularity queries."""

    @pytest.mark.parametrize(
        ("today", "sunday", "saturday"),
        [
            (date(2026, 8, 17), date(2026, 8, 9), date(2026, 8, 15)),  # Monday
            (date(2026, 8, 16), date(2026, 8, 9), date(2026, 8, 15)),  # Sunday: week ended -1d
            (date(2026, 8, 15), date(2026, 8, 2), date(2026, 8, 8)),  # Saturday: week incomplete
            (date(2026, 8, 22), date(2026, 8, 9), date(2026, 8, 15)),  # next Saturday
        ],
    )
    def test_alignment(self, today: date, sunday: date, saturday: date) -> None:
        """Test the returned week starts on Sunday and ends at least 1 day ago."""
        assert popularity.latest_complete_week(today) == (sunday, saturday)
        assert sunday.weekday() == 6
        assert saturday.weekday() == 5


class TestBuildPopularity:
    """build_popularity against a mocked v1 API."""

    def test_matches_terms_unfiltered_and_sorts(self, pem: str, httpx_mock: HTTPXMock) -> None:
        """Test matched rows produce the documented columns, sources and sort."""
        _mock_token(httpx_mock)
        httpx_mock.add_response(
            url=POPULARITY_URL,
            json=_popularity_json(
                [
                    _popularity_row("chess", rank=3, p100=80, p5=4),
                    _popularity_row("puzzle", rank=1, p100=95, p5=5),
                    _popularity_row("solitaire", genre="Casino", rank=7, p100=60, p5=3),
                    _popularity_row("unrelated", rank=2, p100=99, p5=5),
                ]
            ),
        )

        frame, notes = popularity.build_popularity(
            _fetch_result(["Chess", "puzzle"], ["PUZZLE", "solitaire"]),
            client_factory=_factory(pem),
            today=TODAY,
        )

        assert list(frame.columns) == [
            "search_term",
            "source",
            "country_or_region",
            "genre",
            "rank_in_genre",
            "search_popularity_1_to_100",
            "search_popularity_1_to_5",
        ]
        assert frame["search_term"].tolist() == ["puzzle", "chess", "solitaire"]
        assert frame["source"].tolist() == ["Both", "Keywords", "Search Terms"]
        assert frame["genre"].tolist() == ["Games", "Games", "Casino"]
        assert any("Popularity week: 2026-08-09" in note for note in notes)

        body = json.loads(httpx_mock.get_requests(url=POPULARITY_URL)[0].content)
        assert "filters" not in body
        assert set(body["fields"]) == {
            "rankInGenre",
            "searchPopularityInGenre",
            "searchPopularity1to100",
            "searchPopularity1to5",
        }
        assert body["timeRange"] == {
            "start": "2026-08-09",
            "end": "2026-08-15",
            "granularity": "WEEKLY_SUN_SAT",
        }

    def test_empty_week_falls_back_one_week(self, pem: str, httpx_mock: HTTPXMock) -> None:
        """Test an empty latest week retries exactly one week earlier."""
        _mock_token(httpx_mock)
        httpx_mock.add_response(url=POPULARITY_URL, json=_popularity_json([]))
        httpx_mock.add_response(
            url=POPULARITY_URL, json=_popularity_json([_popularity_row("chess")])
        )

        frame, notes = popularity.build_popularity(
            _fetch_result(["chess"], []),
            client_factory=_factory(pem),
            today=TODAY,
        )

        requests = httpx_mock.get_requests(url=POPULARITY_URL)
        assert len(requests) == 2
        first = json.loads(requests[0].content)["timeRange"]
        second = json.loads(requests[1].content)["timeRange"]
        assert date.fromisoformat(second["start"]) == (
            date.fromisoformat(first["start"]) - timedelta(days=7)
        )
        assert second == {
            "start": "2026-08-02",
            "end": "2026-08-08",
            "granularity": "WEEKLY_SUN_SAT",
        }
        assert frame["search_term"].tolist() == ["chess"]
        assert any("Popularity week: 2026-08-02" in note for note in notes)

    def test_no_matches_adds_popularity_floor_note(self, pem: str, httpx_mock: HTTPXMock) -> None:
        """Test terms absent from both weeks yield an empty frame and note."""
        _mock_token(httpx_mock)
        httpx_mock.add_response(
            url=POPULARITY_URL,
            json=_popularity_json([_popularity_row("unrelated")]),
        )

        frame, notes = popularity.build_popularity(
            _fetch_result(["obscure keyword"], []),
            client_factory=_factory(pem),
            today=TODAY,
        )

        assert frame.empty
        assert any("popularity floor" in note for note in notes)

    def test_no_terms_skips_without_network(self, pem: str, httpx_mock: HTTPXMock) -> None:
        """Test empty keyword/search-term frames skip the query entirely."""
        frame, notes = popularity.build_popularity(
            _fetch_result([], []),
            client_factory=_factory(pem),
            today=TODAY,
        )

        assert frame.empty
        assert notes
        assert not httpx_mock.get_requests()

    def test_auto_selects_account_matching_org_id(
        self, pem: str, httpx_mock: HTTPXMock, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test the ACL ad account whose ID equals ASA_ORG_ID is selected."""
        monkeypatch.setenv("ASA_ORG_ID", "2957360")
        _mock_token(httpx_mock)
        httpx_mock.add_response(
            url=ACLS_URL,
            json={
                "result": {
                    "acls": [
                        {"adAccount": {"id": 111, "name": "Other", "orgId": 1}, "roles": ["Admin"]},
                        {
                            "adAccount": {"id": 2957360, "name": "Main", "orgId": 2957360},
                            "roles": ["Admin"],
                        },
                    ]
                }
            },
        )
        httpx_mock.add_response(
            url=POPULARITY_URL, json=_popularity_json([_popularity_row("chess")])
        )

        frame, notes = popularity.build_popularity(
            _fetch_result(["chess"], []),
            client_factory=_factory(pem, ad_account_id=None),
            today=TODAY,
        )

        request = httpx_mock.get_requests(url=POPULARITY_URL)[0]
        assert request.headers["X-AP-Context"] == "adAccountId=2957360"
        assert not frame.empty
        assert any("2957360" in note for note in notes)

    def test_ambiguous_accounts_skip_with_note(
        self, pem: str, httpx_mock: HTTPXMock, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test multiple non-matching ACL accounts degrade to a skip note."""
        monkeypatch.delenv("ASA_ORG_ID", raising=False)
        _mock_token(httpx_mock)
        httpx_mock.add_response(
            url=ACLS_URL,
            json={
                "result": {
                    "acls": [
                        {"adAccount": {"id": 111, "name": "A", "orgId": 1}, "roles": ["Admin"]},
                        {"adAccount": {"id": 222, "name": "B", "orgId": 2}, "roles": ["Admin"]},
                    ]
                }
            },
        )

        frame, notes = popularity.build_popularity(
            _fetch_result(["chess"], []),
            client_factory=_factory(pem, ad_account_id=None),
            today=TODAY,
        )

        assert frame.empty
        assert any("ad account" in note.lower() for note in notes)
        assert not httpx_mock.get_requests(url=POPULARITY_URL)

    def test_api_error_returns_note_without_raising(self, pem: str, httpx_mock: HTTPXMock) -> None:
        """Test an API error yields an empty frame plus explanatory note."""
        _mock_token(httpx_mock)
        httpx_mock.add_response(
            url=POPULARITY_URL,
            status_code=403,
            json={"error": {"code": "FORBIDDEN", "message": "feature not enabled"}},
        )

        frame, notes = popularity.build_popularity(
            _fetch_result(["chess"], []),
            client_factory=_factory(pem),
            today=TODAY,
        )

        assert frame.empty
        assert any("feature not enabled" in note for note in notes)


def _mock_v5_api(httpx_mock: HTTPXMock) -> None:
    """Mock the v5 endpoints the analyze fetch pipeline hits."""
    httpx_mock.add_response(url=TOKEN_URL, json=token_json(), is_reusable=True)
    httpx_mock.add_response(
        url=f"{API}/campaigns?limit=1000&offset=0", json=campaigns_json(), is_reusable=True
    )
    httpx_mock.add_response(
        url=re.compile(rf"{API}/search/apps\?.*"), json={"data": []}, is_reusable=True
    )
    rows = [
        report_row(
            {"campaignId": 1, "campaignName": "Campaign One"},
            [("2026-07-01", 1000, 100, 10, "50.0")],
        )
    ]
    httpx_mock.add_response(
        url=f"{API}/reports/campaigns", json=report_json(rows), is_reusable=True
    )
    for cid in (1, 2):
        for tail in ("adgroups", "keywords", "searchterms", "ads"):
            httpx_mock.add_response(
                url=f"{API}/reports/campaigns/{cid}/{tail}",
                json=report_json(rows),
                is_reusable=True,
            )


@pytest.mark.httpx_mock(assert_all_responses_were_requested=False)
class TestAnalyzeWiring:
    """asa analyze integration with the popularity sheet."""

    @pytest.mark.usefixtures("asa_env")
    def test_popularity_failure_warns_and_completes(
        self, httpx_mock: HTTPXMock, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test a raising build_popularity degrades to a warning line."""

        def _boom(_result: fetch.FetchResult, **_kwargs: object) -> object:
            """Simulate an unexpected popularity failure."""
            raise RuntimeError("popularity exploded")

        monkeypatch.setattr(popularity, "build_popularity", _boom)
        _mock_v5_api(httpx_mock)
        out = tmp_path / "report.xlsx"

        result = runner.invoke(app, ["analyze", "--output", str(out)])

        assert result.exit_code == 0, result.output
        combined = result.output + (result.stderr or "")
        assert "Warning:" in combined
        assert "popularity exploded" in combined
        assert out.exists()
        assert "Search Popularity" not in load_workbook(out).sheetnames

    @pytest.mark.usefixtures("asa_env")
    def test_popularity_sheet_written_when_data_present(
        self, httpx_mock: HTTPXMock, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test a populated popularity frame lands on its own sheet."""
        frame = pd.DataFrame(
            [
                {
                    "search_term": "chess",
                    "source": "Keywords",
                    "country_or_region": "US",
                    "genre": "Games",
                    "rank_in_genre": 3,
                    "search_popularity_1_to_100": 80,
                    "search_popularity_1_to_5": 4,
                }
            ]
        )

        def _fake(_result: fetch.FetchResult, **_kwargs: object) -> tuple[pd.DataFrame, list[str]]:
            """Return a canned popularity frame and note."""
            return frame, ["Popularity week: 2026-08-09 – 2026-08-15"]  # noqa: RUF001

        monkeypatch.setattr(popularity, "build_popularity", _fake)
        _mock_v5_api(httpx_mock)
        out = tmp_path / "report.xlsx"

        result = runner.invoke(app, ["analyze", "--output", str(out)])

        assert result.exit_code == 0, result.output
        ws = load_workbook(out)["Search Popularity"]
        assert ws["A1"].value == "Popularity week: 2026-08-09 – 2026-08-15"  # noqa: RUF001
        assert ws["A2"].value == "Search Term"
        assert ws["A3"].value == "chess"
