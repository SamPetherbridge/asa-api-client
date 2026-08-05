"""End-to-end CLI tests with a mocked transport."""

import re
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pytest_httpx import HTTPXMock
from typer.testing import CliRunner

from asa_api_client.cli.analyze import app
from tests.unit.cli.conftest import (
    API,
    TOKEN_URL,
    campaigns_json,
    report_json,
    report_row,
    token_json,
)

runner = CliRunner()


def _mock_api(httpx_mock: HTTPXMock) -> None:
    httpx_mock.add_response(url=TOKEN_URL, json=token_json(), is_reusable=True)
    httpx_mock.add_response(
        url=f"{API}/campaigns?limit=1000&offset=0", json=campaigns_json(), is_reusable=True
    )
    httpx_mock.add_response(
        url=re.compile(rf"{API}/search/apps\?.*"), json={"data": []}, is_reusable=True
    )
    rows = [
        report_row(
            {"campaignId": 1, "campaignName": "Campaign One"},
            [("2026-07-01", 1000, 100, 10, "50.0")],
        )
    ]
    httpx_mock.add_response(
        url=f"{API}/reports/campaigns", json=report_json(rows), is_reusable=True
    )
    for cid in (1, 2):
        for tail in ("adgroups", "keywords", "searchterms", "ads"):
            httpx_mock.add_response(
                url=f"{API}/reports/campaigns/{cid}/{tail}",
                json=report_json(rows),
                is_reusable=True,
            )


@pytest.mark.httpx_mock(assert_all_responses_were_requested=False)
class TestAnalyzeCommand:
    """asa analyze end to end."""

    @pytest.mark.usefixtures("asa_env")
    def test_happy_path(self, httpx_mock: HTTPXMock, tmp_path: Path) -> None:
        """Writes the workbook, prints its path and a headline."""
        _mock_api(httpx_mock)
        out = tmp_path / "report.xlsx"
        result = runner.invoke(app, ["analyze", "--output", str(out)])
        assert result.exit_code == 0, result.output
        assert out.exists()
        assert str(out) in result.output
        assert "spend" in result.output and "installs" in result.output
        wb = load_workbook(out)
        assert wb.sheetnames[0] == "Summary"

    @pytest.mark.usefixtures("asa_env")
    def test_app_filter_and_default_name(
        self, httpx_mock: HTTPXMock, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """--app scopes; default filename includes the adam id."""
        _mock_api(httpx_mock)
        monkeypatch.chdir(tmp_path)
        result = runner.invoke(app, ["analyze", "--app", "111"])
        assert result.exit_code == 0, result.output
        produced = list(tmp_path.glob("asa-analysis-111-*.xlsx"))
        assert len(produced) == 1

    @pytest.mark.usefixtures("asa_env")
    def test_invalid_range_fails_before_network(self, httpx_mock: HTTPXMock) -> None:
        """--from after --to exits non-zero with a clean message and no requests."""
        result = runner.invoke(app, ["analyze", "--from", "2026-05-02", "--to", "2026-05-01"])
        assert result.exit_code != 0
        assert "before" in (result.output + str(result.exception or ""))
        assert not httpx_mock.get_requests()

    def test_missing_credentials_clean_error(
        self, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
    ) -> None:
        """No env vars → one-line configuration error, no traceback."""
        monkeypatch.chdir(tmp_path)  # hermetic: don't pick up a stray repo .env
        for var in (
            "ASA_CLIENT_ID",
            "ASA_TEAM_ID",
            "ASA_KEY_ID",
            "ASA_ORG_ID",
            "ASA_PRIVATE_KEY",
            "ASA_PRIVATE_KEY_PATH",
        ):
            monkeypatch.delenv(var, raising=False)
        result = runner.invoke(app, ["analyze"])
        assert result.exit_code == 1
        assert result.exception is None or isinstance(result.exception, SystemExit)

    @pytest.mark.usefixtures("asa_env")
    def test_unwritable_output_clean_error(self, httpx_mock: HTTPXMock, tmp_path: Path) -> None:
        """A bad --output path fails cleanly, not with a raw traceback."""
        _mock_api(httpx_mock)
        out = tmp_path / "does-not-exist" / "report.xlsx"
        result = runner.invoke(app, ["analyze", "--output", str(out)])
        assert result.exit_code == 1
        assert result.exception is None or isinstance(result.exception, SystemExit)
