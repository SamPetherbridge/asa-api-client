"""Tests for xlsxwriter workbook rendering (read back with openpyxl)."""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from asa_api_client.cli.metrics import LEVEL_KEYS, aggregate, normalize
from asa_api_client.cli.workbook import SummaryData, write_workbook


def _campaign_daily() -> pd.DataFrame:
    return normalize(
        pd.DataFrame(
            [
                {"date": "2026-07-01", "campaign_id": 1, "campaign_name": "One",
                 "campaign_status": "ENABLED", "app_name": "App A",
                 "impressions": 1000, "taps": 100, "total_installs": 10,
                 "local_spend": "50.0"},
                {"date": "2026-07-02", "campaign_id": 2, "campaign_name": "Two",
                 "campaign_status": "ENABLED", "app_name": "App B",
                 "impressions": 500, "taps": 5, "total_installs": 0,
                 "local_spend": "25.0"},
            ]
        )
    )


def _keyword_agg() -> pd.DataFrame:
    return aggregate(
        normalize(
            pd.DataFrame(
                [
                    {"campaign_id": 1, "ad_group_id": 1, "keyword_id": 11,
                     "keyword": "good", "campaign_name": "One", "ad_group_name": "AG",
                     "impressions": 100, "taps": 10, "total_installs": 5,
                     "local_spend": "10"},
                    {"campaign_id": 1, "ad_group_id": 1, "keyword_id": 12,
                     "keyword": "bad", "campaign_name": "One", "ad_group_name": "AG",
                     "impressions": 100, "taps": 10, "total_installs": 0,
                     "local_spend": "30"},
                ]
            )
        ),
        LEVEL_KEYS["keywords"],
    )


def _write(tmp_path: Path) -> Path:
    daily = _campaign_daily()
    kw = _keyword_agg()
    camp_agg = aggregate(daily, LEVEL_KEYS["campaigns"])
    empty = pd.DataFrame()
    summary = SummaryData(
        title="App A · Performance analysis",
        period_label="2026-07-01 – 2026-07-02 (2 days)",  # noqa: RUF001
        timezone="UTC",
        generated_at=datetime(2026, 8, 5, 9, 30),
        kpis={"spend": 75.0, "impressions": 1500.0, "taps": 105.0, "installs": 10.0,
              "ttr": 0.07, "cvr": 0.095, "cpt": 0.714, "cpa": 7.5},
        prior_kpis={"spend": 50.0, "impressions": 1000.0, "taps": 100.0, "installs": 20.0,
                    "ttr": 0.1, "cvr": 0.2, "cpt": 0.5, "cpa": 2.5},
        daily=pd.DataFrame(
            {"date": ["2026-07-01", "2026-07-02"], "spend": [50.0, 25.0], "installs": [10, 0]}
        ),
        per_app=pd.DataFrame(
            {"app_name": ["App A", "App B"], "spend": [50.0, 25.0],
             "impressions": [1000, 500], "taps": [100, 5], "installs": [10, 0]}
        ),
        top_keywords=kw.head(5),
        wasted=kw[(kw["spend"] > 0) & (kw["installs"] == 0)],
    )
    path = tmp_path / "out.xlsx"
    write_workbook(
        path,
        summary=summary,
        analysis={"campaigns": camp_agg, "ad_groups": empty, "keywords": kw,
                  "search_terms": empty, "ads": empty},
        daily={"campaigns": daily, "ad_groups": empty, "keywords": empty,
               "search_terms": empty, "ads": empty},
        notes={"campaigns": [], "ad_groups": [], "keywords": [],
               "search_terms": ["Search terms limited to trailing 90 days."], "ads": []},
    )
    return path


class TestWorkbook:
    """End-to-end workbook structure assertions."""

    def test_sheet_order(self, tmp_path: Path) -> None:
        """Summary, five analysis sheets, then grey daily sheets."""
        wb = load_workbook(_write(tmp_path))
        expected = ["Summary", "Campaigns", "Ad Groups", "Keywords", "Search Terms",
                    "Ads", "Daily · Campaigns", "Daily · Ad Groups", "Daily · Keywords",
                    "Daily · Search Terms", "Daily · Ads"]
        visible = [n for n in wb.sheetnames if n != "Chart Data"]
        assert visible == expected

    def test_summary_content(self, tmp_path: Path) -> None:
        """Title, KPI band with deltas, callout headers."""
        ws = load_workbook(_write(tmp_path))["Summary"]
        assert ws["A1"].value == "App A · Performance analysis"
        assert ws["A4"].value == "Spend"
        assert ws["A5"].value == 75.0
        assert "▲" in ws["A6"].value  # spend up
        assert "▲" in ws["H6"].value and ws["H4"].value == "CPA"  # CPA up = bad arrow up
        texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
        assert "Top 5 keywords by installs" in texts
        assert "Wasted spend" in texts

    def test_currency_and_percent_formats(self, tmp_path: Path) -> None:
        """Spend uses the currency format; TTR uses a percent format."""
        ws = load_workbook(_write(tmp_path))["Summary"]
        assert "$" in ws["A5"].number_format
        assert "%" in ws["E5"].number_format

    def test_analysis_sheet(self, tmp_path: Path) -> None:
        """Campaigns sheet has headers, data, freeze panes and autofilter."""
        ws = load_workbook(_write(tmp_path))["Campaigns"]
        headers = [c.value for c in ws[1]]
        assert "Campaign" in headers and "Spend" in headers and "CPA" in headers
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None
        spend_col = headers.index("Spend") + 1
        assert "$" in ws.cell(row=2, column=spend_col).number_format

    def test_dash_for_nan_cpa(self, tmp_path: Path) -> None:
        """Campaign 2 (0 installs) renders CPA as an em dash."""
        ws = load_workbook(_write(tmp_path))["Campaigns"]
        headers = [c.value for c in ws[1]]
        cpa_col = headers.index("CPA") + 1
        values = {ws.cell(row=r, column=cpa_col).value for r in (2, 3)}
        assert "—" in values

    def test_daily_sheet_tab_color_and_note(self, tmp_path: Path) -> None:
        """Daily tabs are grey; search-term note appears above the table."""
        wb = load_workbook(_write(tmp_path))
        daily = wb["Daily · Campaigns"]
        assert daily.sheet_properties.tabColor is not None
        st = wb["Daily · Search Terms"]
        assert st["A1"].value == "Search terms limited to trailing 90 days."

    def test_chart_data_hidden(self, tmp_path: Path) -> None:
        """Chart Data sheet exists, is hidden, and holds the series."""
        wb = load_workbook(_write(tmp_path))
        cd = wb["Chart Data"]
        assert cd.sheet_state == "hidden"
        assert cd["A1"].value == "date" and cd["B1"].value == "spend"
